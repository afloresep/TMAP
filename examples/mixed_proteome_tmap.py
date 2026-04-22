"""Mixed-proteome TMAP: ESM-2 embeddings across four taxonomic-compartment
classes (α-proteobacteria, cyanobacteria, eukaryotic mitochondrial, eukaryotic
cytosolic), ~7k proteins, mapped with TMAP2.

Demonstrates that TMAP2 produces a coherent tree that cleanly separates these
four classes in seconds on a laptop — the kind of scale/metric-agnostic
exploration the tree layout is built for. Outputs a static PNG, an interactive
HTML, and a quantitative purity report.

Two-stage workflow:

1. Export the protein dataset as FASTA + metadata TSV:
       python examples/mixed_proteome_tmap.py --export

   This downloads ~4 UniProt proteomes (~15 MB) + MitoCarta 3.0 xls (~1 MB)
   and writes:
       examples/data/endosymbiosis/dataset.fasta
       examples/data/endosymbiosis/dataset_metadata.tsv

2. Run ESM-2 on the FASTA externally (any GPU pipeline works). The expected
   output is a single .npz file containing:
       embeddings  (N, 1280) float32   # mean-pooled over residues
       accessions  (N,) object          # must match dataset.fasta order

   Example using fair-esm on a local GPU:
       pip install fair-esm
       esm-extract esm2_t33_650M_UR50D dataset.fasta esm_out/ \\
           --repr_layers 33 --include mean --truncation_seq_length 1022
       python -c "
   import numpy as np, torch
   from pathlib import Path
   from tmap.utils.proteins import read_fasta
   ids, _ = read_fasta('examples/data/endosymbiosis/dataset.fasta')
   accs = [i.split()[0] for i in ids]
   out = np.stack([
       torch.load(f'esm_out/{a}.pt')['mean_representations'][33].numpy()
       for a in accs
   ]).astype('float32')
   np.savez('examples/data/endosymbiosis/embeddings.npz',
            embeddings=out, accessions=np.array(accs, dtype=object))
   "

3. Fit TMAP + produce figures + validation:
       python examples/mixed_proteome_tmap.py \\
           --embeddings examples/data/endosymbiosis/embeddings.npz
       python examples/mixed_proteome_tmap.py \\
           --embeddings examples/data/endosymbiosis/embeddings.npz --validate

Requirements:
    pip install openpyxl matplotlib 'xlrd<2.0'

No torch, no fair-esm — ESM-2 inference happens in the user's own environment.
"""
from __future__ import annotations

import argparse
import csv
import urllib.parse
import urllib.request
from collections import deque
from dataclasses import dataclass
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from numpy.typing import NDArray

from tmap import TMAP
from tmap.utils.proteins import fetch_uniprot, read_fasta

HERE = Path(__file__).parent
DATA_DIR = HERE / "data" / "endosymbiosis"
IMG_DIR = HERE.parent / "paper" / "images"


@dataclass
class ProteinRecord:
    """One protein: identity + metadata; sequence omitted to keep arrays small."""
    accession: str
    organism: str         # e.g. "Homo sapiens"
    # source is one of: "mitocarta", "rickettsia", "pelagibacter",
    # "synechocystis", "yeast-cytosol"
    source: str
    domain: str           # "Bacteria-alpha", "Bacteria-cyano", "Eukarya-mito", "Eukarya-cytosolic"
    compartment: str      # free-form; e.g. "matrix", "cytosol", "-"


UNIPROT_STREAM = "https://rest.uniprot.org/uniprotkb/stream"


def fetch_uniprot_proteome(
    proteome_id: str,
    *,
    cache_dir: Path,
    reviewed_only: bool = False,
) -> tuple[list[str], list[str]]:
    """Fetch a UniProt reference proteome as FASTA, cached to disk.

    Returns (accessions, sequences). Idempotent: subsequent calls read the
    cached file and skip the network.
    """
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_path = cache_dir / f"{proteome_id}.fasta"

    if not cache_path.exists():
        query = f"proteome:{proteome_id}"
        if reviewed_only:
            query += " AND reviewed:true"
        params = urllib.parse.urlencode({
            "query": query,
            "format": "fasta",
            "compressed": "false",
        })
        url = f"{UNIPROT_STREAM}?{params}"
        req = urllib.request.Request(url, headers={"User-Agent": "tmap2/0.1"})
        with urllib.request.urlopen(req, timeout=120) as resp, open(cache_path, "wb") as f:
            f.write(resp.read())

    ids, seqs = read_fasta(cache_path)
    accessions = [_parse_uniprot_accession(h) for h in ids]
    return accessions, list(seqs)


def _parse_uniprot_accession(header: str) -> str:
    """From 'sp|P12345|NAME_SPC' or 'tr|Q9|NAME' return 'P12345' / 'Q9'."""
    parts = header.split("|")
    if len(parts) >= 2 and parts[0] in ("sp", "tr"):
        return parts[1]
    return header.split()[0]


def load_mitocarta(xls_path: Path) -> list[ProteinRecord]:
    """Parse the MitoCarta 3.0 spreadsheet; return ProteinRecord per entry.

    Auto-detects legacy OLE .xls vs OOXML .xlsx by magic bytes and dispatches
    to the right reader. MitoCarta 3.0 is shipped as legacy .xls, which
    requires xlrd (<2.0): ``pip install 'xlrd<2.0'``. OOXML .xlsx files are
    read with openpyxl.

    Note: we read the legacy .xls path directly with xlrd rather than via
    pandas, because modern pandas requires xlrd>=2.0.1, and xlrd 2.0 dropped
    .xls support. Reading directly with xlrd<2.0 sidesteps that conflict.

    Expects at minimum columns: 'UniProt' and ideally
    'MitoCarta3.0_SubMitoLocalization'.
    """
    with open(xls_path, "rb") as f:
        magic = f.read(8)

    def _mk_record(acc_raw, loc_raw) -> ProteinRecord | None:
        if acc_raw is None or acc_raw == "":
            return None
        # UniProt column may list multiple accessions separated by '|'.
        acc = str(acc_raw).split("|")[0].strip()
        if not acc:
            return None
        loc = str(loc_raw) if (loc_raw not in (None, "")) else "-"
        return ProteinRecord(
            accession=acc, organism="Homo sapiens", source="mitocarta",
            domain="Eukarya-mito", compartment=loc,
        )

    out: list[ProteinRecord] = []

    if magic.startswith(b"\xd0\xcf\x11\xe0"):
        # Legacy OLE .xls — read directly with xlrd<2.0.
        import xlrd

        book = xlrd.open_workbook(str(xls_path))
        target_sheet = None
        target_header: list[str] = []
        for sheet in book.sheets():
            if sheet.nrows == 0:
                continue
            header = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
            if "UniProt" in header:
                target_sheet = sheet
                target_header = header
                break
        if target_sheet is None:
            raise ValueError(
                f"No sheet in {xls_path} has a 'UniProt' column. "
                f"Available sheets: {book.sheet_names()}"
            )
        ix_acc = target_header.index("UniProt")
        ix_loc = (target_header.index("MitoCarta3.0_SubMitoLocalization")
                  if "MitoCarta3.0_SubMitoLocalization" in target_header else None)
        for r in range(1, target_sheet.nrows):
            acc_raw = target_sheet.cell_value(r, ix_acc)
            loc_raw = target_sheet.cell_value(r, ix_loc) if ix_loc is not None else None
            rec = _mk_record(acc_raw, loc_raw)
            if rec is not None:
                out.append(rec)
        return out

    if magic.startswith(b"PK"):
        # OOXML .xlsx — read with openpyxl directly. Hand off an in-memory
        # BytesIO so openpyxl skips its filename-extension check (some files
        # are OOXML-under-the-hood but ship with a '.xls' extension).
        import io

        import openpyxl

        buf = io.BytesIO(Path(xls_path).read_bytes())
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        target_name = None
        target_header = []
        for name in wb.sheetnames:
            ws = wb[name]
            try:
                header = [c.value for c in next(ws.iter_rows(max_row=1))]
            except StopIteration:
                continue
            if "UniProt" in header:
                target_name = name
                target_header = header
                break
        if target_name is None:
            raise ValueError(
                f"No sheet in {xls_path} has a 'UniProt' column. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[target_name]
        ix_acc = target_header.index("UniProt")
        ix_loc = (target_header.index("MitoCarta3.0_SubMitoLocalization")
                  if "MitoCarta3.0_SubMitoLocalization" in target_header else None)
        for row in ws.iter_rows(min_row=2, values_only=True):
            acc_raw = row[ix_acc]
            loc_raw = row[ix_loc] if ix_loc is not None else None
            rec = _mk_record(acc_raw, loc_raw)
            if rec is not None:
                out.append(rec)
        return out

    raise ValueError(
        f"{xls_path}: unrecognized file format (magic bytes {magic[:4].hex()}). "
        f"Expected legacy .xls (OLE) or .xlsx (OOXML)."
    )


def filter_cytosolic(accessions: list[str], *, chunk_size: int = 200) -> list[str]:
    """Keep accessions whose UniProt subcellular-location string mentions
    'Cytoplasm' or 'Cytosol' AND does NOT mention 'Mitochondrion',
    'Plastid', or 'Chloroplast'.

    Used to produce a non-organellar control set from a full proteome.
    """
    out: list[str] = []
    for i in range(0, len(accessions), chunk_size):
        chunk = accessions[i : i + chunk_size]
        ann = fetch_uniprot(chunk, fields=("accession", "cc_subcellular_location"))
        texts = ann["cc_subcellular_location"]
        for j, acc in enumerate(chunk):
            t = str(texts[j]).lower() if j < len(texts) else ""
            if not t:
                continue
            has_cyto  = ("cytoplasm" in t) or ("cytosol" in t)
            has_other = any(k in t for k in ("mitochondrion", "plastid", "chloroplast"))
            if has_cyto and not has_other:
                out.append(acc)
    return out


PROTEOME_SOURCES: tuple[tuple[str, str, str, str], ...] = (
    # (UniProt proteome ID, source tag, domain, organism display name)
    ("UP000002480", "rickettsia",   "Bacteria-alpha", "Rickettsia prowazekii"),
    ("UP000000744", "pelagibacter", "Bacteria-alpha", "Pelagibacter ubique"),
    ("UP000001425", "synechocystis","Bacteria-cyano", "Synechocystis sp. PCC6803"),
    ("UP000002311", "yeast-cytosol","Eukarya-cytosolic", "Saccharomyces cerevisiae"),
)

MITOCARTA_URL = "https://personal.broadinstitute.org/scalvo/MitoCarta3.0/Human.MitoCarta3.0.xls"


def _download_if_missing(url: str, dest: Path) -> Path:
    if dest.exists():
        return dest
    dest.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": "tmap2/0.1"})
    with urllib.request.urlopen(req, timeout=120) as resp, open(dest, "wb") as f:
        f.write(resp.read())
    return dest


def build_endosymbiosis_dataset(
    *, cache_dir: Path,
) -> tuple[list[ProteinRecord], list[str]]:
    """Assemble the full set of proteins + sequences for the atlas.

    Returns (records, sequences) parallel lists. Sequences come from UniProt
    FASTA for proteomes, and a separate UniProt fetch for MitoCarta accessions.
    """
    records: list[ProteinRecord] = []
    sequences: list[str] = []

    # Proteomes (bacterial + yeast cytosolic subset).
    for proteome_id, source, domain, organism in PROTEOME_SOURCES:
        ids, seqs = fetch_uniprot_proteome(proteome_id, cache_dir=cache_dir)
        if source == "yeast-cytosol":
            keep = set(filter_cytosolic(ids))
            keep_mask = [i in keep for i in ids]
            ids = [i for i, k in zip(ids, keep_mask, strict=True) if k]
            seqs = [s for s, k in zip(seqs, keep_mask, strict=True) if k]
        for acc, seq in zip(ids, seqs, strict=True):
            records.append(ProteinRecord(
                accession=acc, organism=organism, source=source,
                domain=domain, compartment="-",
            ))
            sequences.append(seq)

    # MitoCarta — parse the xls for accessions + compartments, then fetch sequences.
    mito_xls = _download_if_missing(MITOCARTA_URL, cache_dir / "Human.MitoCarta3.0.xls")
    mito_records = load_mitocarta(mito_xls)
    mito_accs = [r.accession for r in mito_records]
    ann = fetch_uniprot(mito_accs, fields=("accession", "sequence"))
    seq_by_acc = {
        str(ann["accession"][i]): str(ann["sequence"][i])
        for i in range(len(mito_accs))
    }
    for r in mito_records:
        seq = seq_by_acc.get(r.accession, "")
        if seq:
            records.append(r)
            sequences.append(seq)

    return records, sequences


def write_dataset_fasta(
    records: list[ProteinRecord],
    sequences: list[str],
    *,
    fasta_path: Path,
) -> None:
    """Write records + sequences as a bare FASTA (accession-only headers)."""
    fasta_path.parent.mkdir(parents=True, exist_ok=True)
    with open(fasta_path, "w") as f:
        for r, seq in zip(records, sequences, strict=True):
            f.write(f">{r.accession}\n{seq}\n")


def write_dataset_metadata_tsv(
    records: list[ProteinRecord],
    *,
    tsv_path: Path,
) -> None:
    """Write a TSV keyed by accession with the columns ESM pipelines don't need
    but TMAP analysis does."""
    tsv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(tsv_path, "w") as f:
        f.write("accession\torganism\tsource\tdomain\tcompartment\n")
        for r in records:
            f.write(f"{r.accession}\t{r.organism}\t{r.source}\t{r.domain}\t{r.compartment}\n")


def load_external_embeddings(
    npz_path: Path,
    *,
    expected_accessions: list[str],
) -> tuple[NDArray[np.float32], NDArray]:
    """Load a .npz containing arrays `embeddings` (N, D) and `accessions` (N,).

    Validates that the embedding accession order matches expected_accessions.
    This catches the common "I re-ordered my FASTA" mistake before it poisons
    the tree.
    """
    data = np.load(npz_path, allow_pickle=True)
    if "embeddings" not in data.files or "accessions" not in data.files:
        raise KeyError(
            f"{npz_path} must contain 'embeddings' and 'accessions' arrays; "
            f"found {list(data.files)}."
        )
    accs = list(data["accessions"])
    if accs != list(expected_accessions):
        first_diff = next(
            (i for i, (a, b) in enumerate(zip(accs, expected_accessions)) if a != b),
            min(len(accs), len(expected_accessions)),
        )
        got = accs[first_diff] if first_diff < len(accs) else "<end>"
        want = (expected_accessions[first_diff]
                if first_diff < len(expected_accessions) else "<end>")
        raise ValueError(
            f"{npz_path} accession order does not match the FASTA. "
            f"First divergence at index {first_diff}: "
            f"got {got}, expected {want}."
        )
    X = np.asarray(data["embeddings"], dtype=np.float32)
    return X, np.array(accs, dtype=object)


def _multi_source_bfs_hops(tree, source_mask: NDArray) -> NDArray:
    """Hop count from every node in the tree to its nearest source (nodes where
    source_mask is True). Multi-source BFS; returns int32 array of shape
    (tree.n_nodes,) with a large sentinel for unreachable nodes.
    """
    n = tree.n_nodes
    hops = np.full(n, np.iinfo(np.int32).max, dtype=np.int32)
    queue: deque = deque()
    for i in np.where(source_mask)[0]:
        hops[int(i)] = 0
        queue.append(int(i))
    while queue:
        node = queue.popleft()
        for neighbor, _weight in tree.neighbors(node):
            if hops[neighbor] > hops[node] + 1:
                hops[neighbor] = hops[node] + 1
                queue.append(int(neighbor))
    return hops


def mito_to_alpha_path_stats(
    *, tree, sources: NDArray,
) -> dict[str, float]:
    """For each mitocarta protein, measure the shortest tree path (in hops)
    to the nearest α-proteobacterial protein and to the nearest yeast-cytosolic
    protein. Return medians and the mito-sample count.

    Implementation: two multi-source BFS passes, O(N) each, instead of
    O(M × A) individual tree.path() calls.
    """
    mito_idx  = np.where(sources == "mitocarta")[0]
    alpha_mask = (sources == "rickettsia") | (sources == "pelagibacter")
    cyto_mask  = sources == "yeast-cytosol"

    hops_alpha_all = _multi_source_bfs_hops(tree, alpha_mask)
    hops_cyto_all  = _multi_source_bfs_hops(tree, cyto_mask) if cyto_mask.any() else None

    hops_alpha = hops_alpha_all[mito_idx]
    hops_cyto  = hops_cyto_all[mito_idx] if hops_cyto_all is not None else np.array([])

    # Filter out the unreachable sentinel values.
    INF = np.iinfo(np.int32).max
    finite_alpha = hops_alpha[hops_alpha < INF]
    finite_cyto  = hops_cyto[hops_cyto < INF]

    med_alpha = float(np.median(finite_alpha)) if finite_alpha.size else float("nan")
    med_cyto  = float(np.median(finite_cyto))  if finite_cyto.size  else float("nan")
    return {
        "median_hops_to_alpha":   med_alpha,
        "median_hops_to_cytosol": med_cyto,
        "n_mito":                 float(len(mito_idx)),
    }


def alpha_branch_mito_fraction(
    *, tree, sources: NDArray, radius: int = 2,
) -> float:
    """Fraction of mitocarta proteins within `radius` tree hops of any
    α-proteobacterial protein.

    Implementation: one multi-source BFS, O(N), instead of M × A tree.path calls.
    """
    mito_idx  = np.where(sources == "mitocarta")[0]
    alpha_mask = (sources == "rickettsia") | (sources == "pelagibacter")
    if mito_idx.size == 0 or not alpha_mask.any():
        return 0.0
    hops_to_alpha = _multi_source_bfs_hops(tree, alpha_mask)
    mito_hops = hops_to_alpha[mito_idx]
    return float((mito_hops <= radius).sum()) / float(len(mito_idx))


DOMAIN_PALETTE = {
    "Bacteria-alpha":    "#d62728",  # red
    "Bacteria-cyano":    "#2ca02c",  # green
    "Eukarya-mito":      "#ff7f0e",  # orange
    "Eukarya-cytosolic": "#1f77b4",  # blue
}


def plot_endosymbiosis_tree(
    *,
    layout: NDArray,
    domains: NDArray,
    tree_edges,
    out_path: Path,
    title: str = "Endosymbiosis TMAP: ESM-2 sequence embeddings",
) -> None:
    """Full tree colored by domain."""
    fig, ax = plt.subplots(figsize=(8.5, 7.5), dpi=150)

    # Plot edges in light gray first so points sit on top.
    for u, v in tree_edges:
        ax.plot([layout[u, 0], layout[v, 0]],
                [layout[u, 1], layout[v, 1]],
                "-", lw=0.35, color="#cccccc", zorder=1)

    for domain, color in DOMAIN_PALETTE.items():
        mask = domains == domain
        if not mask.any():
            continue
        ax.scatter(layout[mask, 0], layout[mask, 1],
                   s=8, c=color, linewidths=0, alpha=0.85,
                   label=f"{domain} (n={int(mask.sum())})", zorder=2)

    ax.set_title(title, fontsize=11)
    ax.set_xticks([])
    ax.set_yticks([])
    ax.set_aspect("equal", adjustable="datalim")
    ax.legend(loc="lower left", fontsize=8, frameon=False)

    fig.tight_layout()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, bbox_inches="tight")
    plt.close(fig)


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Endosymbiosis mitochondrial TMAP.")
    p.add_argument("--cache-dir", type=Path, default=DATA_DIR)
    p.add_argument("--export", action="store_true",
                   help="Assemble the dataset and write FASTA + metadata TSV for "
                        "external ESM-2 embedding, then exit.")
    p.add_argument("--embeddings", type=Path, default=None,
                   help="Path to the .npz containing user-produced ESM-2 embeddings. "
                        "Required unless --export is given.")
    p.add_argument("--n-neighbors", type=int, default=20)
    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--validate", action="store_true")
    p.add_argument("--max-cells", type=int, default=0,
                   help="If >0, subsample to this many proteins (stratified by source).")
    return p


def _stratified_subsample(
    records, sequences, sources, domains, accessions, *, max_cells: int, seed: int,
):
    """Uniformly stratified subsampling by source; returns the same 5 lists/arrays."""
    rng = np.random.default_rng(seed)
    total = len(records)
    keep: list[int] = []
    for src in sorted(set(sources.tolist())):
        idx = np.where(sources == src)[0]
        take = min(len(idx), max(1, int(max_cells * len(idx) / total)))
        keep.extend(rng.choice(idx, size=take, replace=False).tolist())
    keep = sorted(keep)
    return (
        [records[i]   for i in keep],
        [sequences[i] for i in keep],
        sources[keep],
        domains[keep],
        [accessions[i] for i in keep],
    )


def main() -> None:
    args = _build_parser().parse_args()

    # Stage A: --export produces FASTA + metadata TSV and exits.
    if args.export:
        print("Building dataset …")
        records, sequences = build_endosymbiosis_dataset(cache_dir=args.cache_dir)
        sources    = np.array([r.source for r in records])
        domains    = np.array([r.domain for r in records])
        accessions = [r.accession for r in records]

        if args.max_cells and len(records) > args.max_cells:
            records, sequences, sources, domains, accessions = _stratified_subsample(
                records, sequences, sources, domains, accessions,
                max_cells=args.max_cells, seed=args.seed,
            )

        counts_by_source = {s: int((sources == s).sum()) for s in sorted(set(sources.tolist()))}
        print(f"  Total: {len(records):,}")
        for k, v in counts_by_source.items():
            print(f"    {k}: {v:,}")

        fasta_path = args.cache_dir / "dataset.fasta"
        tsv_path   = args.cache_dir / "dataset_metadata.tsv"
        write_dataset_fasta(records, sequences, fasta_path=fasta_path)
        write_dataset_metadata_tsv(records, tsv_path=tsv_path)
        print(f"  Wrote {fasta_path}")
        print(f"  Wrote {tsv_path}")
        print("\nNext: run your ESM-2 pipeline on the FASTA and save a .npz "
              "containing arrays 'embeddings' (N,1280) and 'accessions' (N,) "
              "in the same order as the FASTA. Then rerun this script with "
              "--embeddings <path.npz>. See the module docstring for a ready-to-paste "
              "fair-esm recipe.")
        return

    # Stage B requires --embeddings.
    if args.embeddings is None:
        raise SystemExit(
            "error: either --export (to write FASTA+TSV) or "
            "--embeddings <path.npz> (to fit TMAP) is required."
        )

    IMG_DIR.mkdir(parents=True, exist_ok=True)

    fasta_path = args.cache_dir / "dataset.fasta"
    tsv_path   = args.cache_dir / "dataset_metadata.tsv"
    if not fasta_path.exists() or not tsv_path.exists():
        raise SystemExit(
            f"Dataset files missing. Run --export first to produce {fasta_path} "
            f"and {tsv_path}."
        )

    ids, _seqs = read_fasta(fasta_path)
    accessions = [i.split()[0] for i in ids]

    # Reload metadata keyed by accession.
    meta_by_acc: dict[str, dict[str, str]] = {}
    with open(tsv_path) as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            meta_by_acc[row["accession"]] = row
    sources = np.array([meta_by_acc[a]["source"] for a in accessions])
    domains = np.array([meta_by_acc[a]["domain"] for a in accessions])

    print(f"Loading embeddings from {args.embeddings} …")
    X, _ = load_external_embeddings(args.embeddings, expected_accessions=accessions)
    print(f"  Embeddings: {X.shape}")

    print(f"Fitting TMAP (metric=cosine, n_neighbors={args.n_neighbors}) …")
    model = TMAP(
        metric="cosine", n_neighbors=args.n_neighbors, seed=args.seed, store_index=True,
    ).fit(X.astype(np.float32, copy=False))

    tree = model.tree_
    layout = model.embedding_
    tree_edges = [(int(u), int(v)) for (u, v) in tree.edges]

    plot_endosymbiosis_tree(
        layout=layout, domains=domains, tree_edges=tree_edges,
        out_path=IMG_DIR / "mixed_proteome_tree.png",
    )
    print("  Figure written: mixed_proteome_tree.png")

    # Interactive HTML for exploration.
    html_path = IMG_DIR / "mixed_proteome_tree.html"
    viz = model.to_tmapviz()
    viz.title = f"Mixed proteome TMAP ({X.shape[0]:,} proteins)"
    viz.add_color_layout("domain", domains.tolist(), categorical=True, color="tab10")
    viz.add_color_layout("source", sources.tolist(), categorical=True, color="tab20")
    viz.add_label("accession", accessions)
    viz.write_html(html_path)
    print(f"  Interactive HTML: {html_path}")

    # Domain-purity stats for the coherent-branches claim.
    from tmap.graph.analysis import boundary_edges, subtree_purity
    n_edges = len(tree.edges)
    boundaries = boundary_edges(tree, domains)
    n_cross = len(boundaries)
    n_same = n_edges - n_cross
    same_frac = n_same / n_edges if n_edges else 0.0

    purity = subtree_purity(tree, domains, min_size=100)
    valid_purity = purity[~np.isnan(purity)] if purity.size else np.array([])

    print(f"  tree edges: {n_edges:,}")
    print(f"  same-domain edges: {n_same:,} ({same_frac:.1%})")
    print(f"  cross-domain edges: {n_cross:,} ({1 - same_frac:.1%})")
    if valid_purity.size:
        print(f"  subtree purity (min_size=100): "
              f"mean={valid_purity.mean():.3f} "
              f"median={float(np.median(valid_purity)):.3f}")

    if args.validate:
        print("\nValidation:")
        criteria = [
            (
                "same-domain edge fraction >= 0.90 (tree is coherent by domain)",
                same_frac >= 0.90,
                f"same_frac={same_frac:.3f}",
            ),
            (
                "at least 4 subtrees (min_size=100) with purity >= 0.95 (one per domain)",
                valid_purity.size >= 4 and int((valid_purity >= 0.95).sum()) >= 4,
                f"n_subtrees_purity_above_0.95={int((valid_purity >= 0.95).sum())}",
            ),
        ]
        fails = []
        for name, ok, detail in criteria:
            print(f"  [{'PASS' if ok else 'FAIL'}] {name}: {detail}")
            if not ok:
                fails.append(name)
        if fails:
            raise SystemExit(1)
        print(f"\nAll {len(criteria)} criteria passed.")


if __name__ == "__main__":
    main()
