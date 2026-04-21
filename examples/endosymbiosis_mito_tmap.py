"""Endosymbiosis TMAP: ESM-2 embeddings of ~15k proteins spanning
α-proteobacteria, cyanobacteria, the human mitochondrial proteome (MitoCarta
3.0), and a eukaryotic cytosolic control.

Tests the hypothesis that in sequence-embedding space the mitochondrial
proteome sits inside the α-proteobacterial clade of the tree.

Two-stage workflow:

1. Export the protein dataset as FASTA + metadata TSV:
       python examples/endosymbiosis_mito_tmap.py --export

   This downloads ~4 UniProt proteomes (~15 MB) + MitoCarta 3.0 xls (~1 MB)
   and writes:
       examples/data/endosymbiosis/dataset.fasta
       examples/data/endosymbiosis/dataset_metadata.tsv

2. Run ESM-2 on the FASTA externally (any GPU pipeline works). The expected
   output is a single .npz file containing:
       embeddings  (N, 1280) float32   # mean-pooled over residues
       accessions  (N,) object          # must match dataset.fasta order

   Example using fair-esm on a local GPU:
       pip install fair-esm
       esm-extract esm2_t33_650M_UR50D dataset.fasta esm_out/ \\
           --repr_layers 33 --include mean --truncation_seq_length 1022
       python -c "
   import numpy as np, torch
   from pathlib import Path
   from tmap.utils.proteins import read_fasta
   ids, _ = read_fasta('examples/data/endosymbiosis/dataset.fasta')
   accs = [i.split()[0] for i in ids]
   out = np.stack([
       torch.load(f'esm_out/{a}.pt')['mean_representations'][33].numpy()
       for a in accs
   ]).astype('float32')
   np.savez('examples/data/endosymbiosis/embeddings.npz',
            embeddings=out, accessions=np.array(accs, dtype=object))
   "

3. Fit TMAP and produce figures + validation:
       python examples/endosymbiosis_mito_tmap.py \\
           --embeddings examples/data/endosymbiosis/embeddings.npz
       python examples/endosymbiosis_mito_tmap.py \\
           --embeddings examples/data/endosymbiosis/embeddings.npz --validate

Requirements:
    pip install openpyxl matplotlib

No torch, no fair-esm — ESM-2 inference happens in the user's own environment.
"""
from __future__ import annotations

import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path

from tmap.utils.proteins import read_fasta

HERE = Path(__file__).parent
DATA_DIR = HERE / "data" / "endosymbiosis"
IMG_DIR = HERE.parent / "paper" / "images"


@dataclass
class ProteinRecord:
    """One protein: identity + metadata; sequence omitted to keep arrays small."""
    accession: str
    organism: str         # e.g. "Homo sapiens"
    # source is one of: "mitocarta", "rickettsia", "pelagibacter",
    # "synechocystis", "yeast-cytosol"
    source: str
    domain: str           # "Bacteria-alpha", "Bacteria-cyano", "Eukarya-mito", "Eukarya-cytosolic"
    compartment: str      # free-form; e.g. "matrix", "cytosol", "-"


UNIPROT_STREAM = "https://rest.uniprot.org/uniprotkb/stream"


def fetch_uniprot_proteome(
    proteome_id: str,
    *,
    cache_dir: Path,
    reviewed_only: bool = False,
) -> tuple[list[str], list[str]]:
    """Fetch a UniProt reference proteome as FASTA, cached to disk.

    Returns (accessions, sequences). Idempotent: subsequent calls read the
    cached file and skip the network.
    """
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_path = cache_dir / f"{proteome_id}.fasta"

    if not cache_path.exists():
        query = f"proteome:{proteome_id}"
        if reviewed_only:
            query += " AND reviewed:true"
        params = urllib.parse.urlencode({
            "query": query,
            "format": "fasta",
            "compressed": "false",
        })
        url = f"{UNIPROT_STREAM}?{params}"
        req = urllib.request.Request(url, headers={"User-Agent": "tmap2/0.1"})
        with urllib.request.urlopen(req, timeout=120) as resp, open(cache_path, "wb") as f:
            f.write(resp.read())

    ids, seqs = read_fasta(cache_path)
    accessions = [_parse_uniprot_accession(h) for h in ids]
    return accessions, list(seqs)


def _parse_uniprot_accession(header: str) -> str:
    """From 'sp|P12345|NAME_SPC' or 'tr|Q9|NAME' return 'P12345' / 'Q9'."""
    parts = header.split("|")
    if len(parts) >= 2 and parts[0] in ("sp", "tr"):
        return parts[1]
    return header.split()[0]


def load_mitocarta(xls_path: Path) -> list[ProteinRecord]:
    """Parse the MitoCarta 3.0 spreadsheet; return ProteinRecord per entry.

    Expects at minimum columns: 'UniProt', 'MitoCarta3.0_SubMitoLocalization'.
    """
    import io

    import openpyxl
    # MitoCarta ships as '.xls' but the file is actually OOXML (xlsx) under the
    # hood. openpyxl's extension check rejects '.xls' — bypass it by handing
    # off an in-memory BytesIO, which skips the filename-based format check.
    buf = io.BytesIO(Path(xls_path).read_bytes())
    wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    # MitoCarta sheet names vary across minor versions; pick the first sheet
    # whose row 1 contains 'UniProt'.
    target = None
    for name in wb.sheetnames:
        ws = wb[name]
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        if "UniProt" in header:
            target = (name, header)
            break
    if target is None:
        raise ValueError(f"No sheet in {xls_path} has a 'UniProt' column.")

    sheet_name, header = target
    ws = wb[sheet_name]
    ix_acc = header.index("UniProt")
    ix_loc = (header.index("MitoCarta3.0_SubMitoLocalization")
              if "MitoCarta3.0_SubMitoLocalization" in header else None)

    out: list[ProteinRecord] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        acc_raw = row[ix_acc]
        if acc_raw in (None, ""):
            continue
        # UniProt column may list multiple accessions separated by '|'.
        acc = str(acc_raw).split("|")[0].strip()
        if not acc:
            continue
        loc = str(row[ix_loc]) if (ix_loc is not None and row[ix_loc] is not None) else "-"
        out.append(ProteinRecord(
            accession=acc, organism="Homo sapiens", source="mitocarta",
            domain="Eukarya-mito", compartment=loc,
        ))
    return out


def main() -> None:
    raise NotImplementedError("Filled in later tasks.")


if __name__ == "__main__":
    main()
